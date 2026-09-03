"""itcen-proposal xlsx 에 시트 추가: 5.AI신기술50 + 6.접목100 + 7.AIRiskLighthouse"""
import sys, os, shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

src = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                   'outputs', 'itcen', 'itcen-proposal-2026-06-01.xlsx')
shutil.copy(src, src + '.bak')
wb = load_workbook(src)

thin = Side(border_style='thin', color='C0C0C0')
brd = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='4472C4')
cat_fill = PatternFill('solid', fgColor='D9E1F2')
cell_font = Font(name='맑은 고딕', size=10)

# Sheet 5: AI 신기술 50
if '5.AI신기술50' in wb.sheetnames:
    del wb['5.AI신기술50']
ws5 = wb.create_sheet('5.AI신기술50')
ws5['A1'] = 'AI 신기술 50개 카탈로그 (2027-2030 메가트렌드)'
ws5['A1'].font = Font(name='맑은 고딕', size=14, bold=True, color='4472C4')
ws5.merge_cells('A1:F1')
hdrs = ['#', '카테고리', '기술명', '한 줄 정의', '핵심 회사-제품', '시점']
for col, h in enumerate(hdrs, 1):
    c = ws5.cell(row=3, column=col, value=h)
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', vertical='center'); c.border = brd
ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 28
ws5.column_dimensions['D'].width = 50
ws5.column_dimensions['E'].width = 35
ws5.column_dimensions['F'].width = 10

TECH_50 = [
    ('A. 추론-인지', [
        (1, 'Reasoning Models', '추론시간 길게 = 정답률↑. 사고 과정 공개', 'OpenAI o3-DeepSeek R1-Claude Extended Thinking', '지금'),
        (2, 'Self-Critique', 'AI 가 자기 답 비판-재검토-수정', 'Reflexion-Self-RAG-Constitutional AI', '지금'),
        (3, 'Causal AI', '상관관계 X 인과관계 추론', 'DoWhy-Causica-Ananke', '2027'),
        (4, 'Chain/Tree/Graph-of-Thought', '사고 과정 분기-평가-합의', 'OpenAI-Anthropic-Google', '지금'),
        (5, 'Neurosymbolic AI', 'LLM + 룰베이스-기호 융합', 'IBM-MIT-DeepMind', '2028'),
    ]),
    ('B. 에이전트', [
        (6, 'Agentic AI', '자율 목표-계획-행동', 'Claude Computer Use-OpenAI Operator', '지금'),
        (7, 'Multi-Agent Systems', '여러 AI 협업-토론', 'CrewAI-AutoGen-LangGraph', '지금'),
        (8, 'Computer-Use / Browser-Use', '화면 보고 클릭-웹 자율 탐색', 'Claude Computer Use', '지금'),
        (9, 'Tool Use / MCP', 'AI 가 외부 도구 호출 표준', 'MCP (Anthropic)', '지금'),
    ]),
    ('C. 학습 패러다임', [
        (10, 'Mixture of Experts (MoE)', '토큰별 전문가 활성, 경량 고성능', 'DeepSeek V3-Llama 4-Mixtral', '지금'),
        (11, 'State Space Models', 'Transformer 대체, 긴 컨텍스트 효율', 'Mamba-Jamba-Striped Hyena', '2028'),
        (12, 'Test-Time Training / Compute', '추론 중 학습-계산', 'o1-o3-Claude', '지금'),
    ]),
    ('D. 생성형 AI', [
        (13, 'Text-to-Video', '텍스트 -> 영상 생성', 'Sora-Veo-Kling-Runway', '지금'),
        (14, 'Code Agents', '코드 자율 생성-수정-테스트', 'Cursor-Devin-Claude Code', '지금'),
        (15, 'Multimodal Native', '텍스트-이미지-음성 통합 학습', 'GPT-4o-Gemini 2.5-Claude', '지금'),
    ]),
    ('E. 피지컬 AI', [
        (16, 'World Models', '영상서 물리법칙 학습', 'NVIDIA Cosmos-Sora-Genie-V-JEPA', '2027'),
        (17, 'Vision-Language-Action (VLA)', '보고 듣고 행동. 로봇용 foundation', 'RT-2-Pi0-OpenVLA-NVIDIA GR00T', '2027'),
        (18, 'Embodied AI / Humanoid', 'LLM + 로봇 결합', 'Figure-Tesla Optimus-1X-Apptronik', '2028'),
        (19, 'Sim-to-Real Transfer', '시뮬 학습 -> 현실 배치', 'NVIDIA Isaac-DeepMind', '2027'),
    ]),
    ('F. 양자 AI', [
        (20, 'Quantum ML (QML)', '양자컴퓨터로 ML 학습', 'IBM Quantum-Google-Azure Quantum', '2028-30'),
        (21, 'Variational Quantum Circuits', '양자-고전 하이브리드 학습', 'Qiskit-PennyLane', '2029'),
    ]),
    ('G. 검색-메모리', [
        (22, 'GraphRAG', '지식그래프 + RAG. 관계 기반 검색', 'Microsoft GraphRAG-LangGraph', '지금'),
        (23, 'Memory Architectures', 'AI 장기-일화 기억', 'MemGPT-Letta-Mem0', '2027'),
    ]),
    ('H. AI 보안', [
        (24, 'Prompt Injection Defense', '프롬프트 공격 방어', 'Lakera-HiddenLayer-Protect AI', '지금'),
        (25, 'Mechanistic Interpretability', 'AI 내부 회로 해석 (안전)', 'Anthropic-Goodfire', '2027'),
        (26, 'Deepfake Detection / C2PA', '음성-영상 위조 탐지-출처 표준', 'Reality Defender-Hive-Pindrop', '지금'),
    ]),
    ('I. 프라이버시 AI', [
        (27, 'Federated Learning', '데이터 안 모으고 모델만 학습', 'NVIDIA FLARE-Flower-Owkin', '2027'),
        (28, 'Confidential Computing', '메모리 사용 중 암호화', 'NVIDIA H100-Intel SGX-AMD SEV', '2027'),
        (29, 'Synthetic Data Generation', '합성 데이터로 학습 (개인정보 없이)', 'Gretel-Mostly AI-Tonic', '지금'),
    ]),
    ('J. 인지-정서', [
        (30, 'Affective Computing / Emotion AI', '감정-스트레스 인식', 'Affectiva-Hume-Realeyes', '2027'),
    ]),
    ('K. 보안 신영역', [
        (31, 'Adversarial ML Defense', '모델 탈취-회피 공격 방어', 'Robust Intelligence-CalypsoAI', '지금'),
        (32, 'AI Workload Protection', 'AI 모델-학습데이터-추론 보호', 'Palo Alto AI-Wiz AI-Aim Security', '지금'),
        (33, 'Non-Human Identity (NHI)', 'AI 에이전트-서비스계정 신원관리', 'Astrix-Oasis-Entro', '2027'),
        (34, 'Cybersecurity Mesh Architecture', 'Gartner 분산 보안 표준', 'Fortinet-Cisco-Palo Alto', '2027'),
        (35, 'DSPM (Data Security Posture)', '데이터 위치-민감도 자동 추적', 'Cyera-Sentra-Varonis', '2027'),
    ]),
    ('L. 인증-생체', [
        (36, 'Behavioral Biometrics', '타이핑-마우스-걸음 상시 인증', 'BioCatch-Nuance-Mastercard', '지금'),
        (37, 'Continuous Authentication', '세션 내 상시 신원 검증', 'Plurilock-Cognito', '2027'),
        (38, 'Passkeys / FIDO2', '비밀번호 종말', 'FIDO Alliance-Apple-Google', '지금'),
    ]),
    ('M. 데이터-학습', [
        (39, 'DPO', 'RLHF 대체, 학습 비용↓', 'Stanford-Anthropic-Meta', '지금'),
        (40, 'RLHF', '인간 피드백 강화학습', 'OpenAI-Anthropic-DeepMind', '지금'),
        (41, 'Constitutional AI', '헌법 따라 자기 규제', 'Anthropic', '지금'),
        (42, 'LoRA / QLoRA', '저비용 fine-tuning', 'Microsoft-Hugging Face', '지금'),
    ]),
    ('N. 검색-인프라', [
        (43, 'Vector Databases', '임베딩 기반 검색 인프라', 'Pinecone-Weaviate-Chroma-Qdrant', '지금'),
        (44, 'HyDE', '가상 답 먼저 만들고 검색', 'LangChain-LlamaIndex', '지금'),
        (45, 'Long Context (1M+)', '1M+ 토큰 컨텍스트', 'Claude-Gemini-Llama', '지금'),
    ]),
    ('O. AI 거버넌스', [
        (46, 'AI Governance Platforms', '라이프사이클-감사-모델 카드', 'IBM watsonx-Credo AI-Holistic AI', '2027'),
        (47, 'Bias Detection / Explainability', '편향 탐지-설명 (LIME-SHAP)', 'Fiddler-Arthur-H2O', '지금'),
    ]),
    ('P. 도메인-신영역', [
        (48, 'Domain Foundation Models', '의료-법무-금융 특화 LLM', 'Med-PaLM-Harvey-BloombergGPT', '지금'),
        (49, 'AI Search Engines', 'LLM 기반 검색', 'Perplexity-SearchGPT-You.com', '지금'),
        (50, 'Ambient Invisible Intelligence', '저비용 IoT + 실시간 추적 (Gartner)', '통합 표준 형성 중', '2027'),
    ]),
]

row = 4
for cat, items in TECH_50:
    cell = ws5.cell(row=row, column=2, value=cat)
    cell.font = Font(name='맑은 고딕', size=10, bold=True, color='4472C4')
    cell.fill = cat_fill
    ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = brd
    row += 1
    for num, name, desc, vendor, when in items:
        ws5.cell(row=row, column=1, value=num).border = brd
        ws5.cell(row=row, column=2, value='').border = brd
        c = ws5.cell(row=row, column=3, value=name)
        c.font = Font(name='맑은 고딕', size=10, bold=True); c.border = brd
        c = ws5.cell(row=row, column=4, value=desc)
        c.font = cell_font; c.alignment = Alignment(wrap_text=True, vertical='center'); c.border = brd
        c = ws5.cell(row=row, column=5, value=vendor)
        c.font = cell_font; c.alignment = Alignment(wrap_text=True, vertical='center'); c.border = brd
        ws5.cell(row=row, column=6, value=when).border = brd
        ws5.row_dimensions[row].height = 24
        row += 1
print(f'Sheet 5 (50 tech): {row-4} rows')

# Sheet 6: 100 접목
if '6.접목100' in wb.sheetnames:
    del wb['6.접목100']
ws6 = wb.create_sheet('6.접목100')
ws6['A1'] = '50 신기술 × ITCEN CORE × 부서 = 100 접목 사업-솔루션'
ws6['A1'].font = Font(name='맑은 고딕', size=14, bold=True, color='4472C4')
ws6.merge_cells('A1:E1')
hdrs6 = ['#', '기반 기술', '접목 아이디어', '활용 자산-부서', '시너지']
for col, h in enumerate(hdrs6, 1):
    c = ws6.cell(row=3, column=col, value=h)
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', vertical='center'); c.border = brd
ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 22
ws6.column_dimensions['C'].width = 60
ws6.column_dimensions['D'].width = 22
ws6.column_dimensions['E'].width = 10

APPS_100 = [
    (1, 'Reasoning Models', '행동위험 점수 추론 모델 (왜 위험한지 사고 과정 공개)', '부서 IP', '★★★'),
    (2, 'Reasoning Models', '내부회계 부정거래 추론 + 근거 자동 보고서', '내부회계', '★★'),
    (3, 'Self-Critique', 'UEBA 위험점수 1차->2차 AI 재검토->합의점수 (차별화 IP)', '부서 핵심', '★★★'),
    (4, 'Self-Critique', 'AI CCTV 이상감지 재검토 -> 알람 정확도 ↑', 'AI CCTV', '★★★'),
    (5, 'Causal AI', '부정행위 인과관계 분석 (단순 상관 X -> 진짜 원인)', '부서-내부회계', '★★★'),
    (6, 'Causal AI', '중대재해 사고 인과분석 + 예방 권고', '건설 ERP', '★★'),
    (7, 'CoT/ToT', '위험분석 의사결정 트리 자동화', 'GRC', '★★'),
    (8, 'Graph-of-Thought', '컴플라이언스 영향분석 (규제 변경 -> 전사 영향)', 'GRC-EPM', '★★'),
    (9, 'Neurosymbolic', '법규 룰 + LLM 통합 GRC 차세대', 'GRC', '★★★'),
    (10, 'Neurosymbolic', '회계 룰베이스 + AI 부정탐지 융합', '내부회계', '★★'),
    (11, 'Agentic AI', '24/7 자율 Risk Officer', '부서', '★★★'),
    (12, 'Agentic AI', '내부회계 자율 감사 에이전트', '내부회계', '★★'),
    (13, 'Multi-Agent', '부서 위험점수 다중 AI 토론 합의', '부서 IP', '★★★'),
    (14, 'Multi-Agent', '카지노 부정 다관점 분석', '카지노 VMS', '★★★'),
    (15, 'Computer-Use', '자율 컴플라이언스 점검', 'GRC', '★★'),
    (16, 'Computer-Use', '자율 보안 모의침투 테스트', '보안', '★★'),
    (17, 'MCP 표준', 'ITCEN CORE 제품 LLM 통합 표준', '전 제품', '★★★'),
    (18, 'MCP 표준', '외부 데이터 소스 자동 연결', '데이터', '★'),
    (19, 'MoE', '도메인별 sLLM (회계-VMS-GRC) 경량 운영', '전 제품', '★★'),
    (20, 'MoE', '다 산업 행동위험 모델', '부서', '★★'),
    (21, 'SSM (Mamba)', '1년치 행동데이터 긴 컨텍스트 분석', '부서-VMS', '★★'),
    (22, 'SSM', 'CCTV 장기간 영상 분석', 'AI CCTV', '★★'),
    (23, 'Test-Time Compute', '고위험 사건 깊은 추론', '부서', '★★'),
    (24, 'Test-Time Training', '신규 위협 즉시 적응', '보안', '★★'),
    (25, 'Text-to-Video', '부정행위 시뮬레이션 학습데이터 생성', '부서', '★★'),
    (26, 'Text-to-Video', '재해-사고 시뮬레이션', '건설', '★★'),
    (27, 'Code Agents', '보안 자동 패치 + 컴플라이언스 자동화', '보안', '★★'),
    (28, 'Code Agents', '컴플라이언스 규칙 자동 코드화 (Rule-as-Code)', 'GRC', '★★★'),
    (29, 'Multimodal Native', 'CCTV + 음성 + 텍스트 통합 행동분석', 'AI CCTV-VMS', '★★★'),
    (30, 'Multimodal', '회의록 + 영상 + 자료 통합 위험감지', '부서', '★★'),
    (31, 'World Models', 'ITCEN CORE 디지털트윈 + Cosmos 자동 학습', '디지털트윈', '★★★'),
    (32, 'World Models', '공장-건설현장 사고 시뮬레이션', '건설 ERP', '★★'),
    (33, 'VLA', '작업자 의도-행동 추론', '부서-CCTV', '★★★'),
    (34, 'VLA', '카지노 딜러-플레이어 행동 분석', '카지노 VMS', '★★★'),
    (35, 'Embodied AI', '보안 로봇 자율 순찰 (NVIDIA GR00T)', '디지털트윈', '★★'),
    (36, 'Embodied AI', '무인 데이터센터 운영', 'ITO', '★'),
    (37, 'Sim-to-Real', '시뮬 학습 -> 실제 CCTV 배치', 'AI CCTV', '★★'),
    (38, 'Sim-to-Real', '작업안전 시뮬 -> 현장 (산안법 의무)', '건설 ERP', '★★'),
    (39, 'Quantum ML', '금융 부정거래 양자최적화 (IBM Quantum 무료)', '내부회계', '★★'),
    (40, 'Quantum ML', '공급망 위험 양자 시뮬', '건설', '★★'),
    (41, 'Variational QC', '행동패턴 양자 분류', '부서', '★★'),
    (42, 'Variational QC', '신용평가 양자 가속', '내부회계', '★'),
    (43, 'GraphRAG', '행동패턴 그래프 + LLM (관계 기반 위험분석)', '부서 IP', '★★★'),
    (44, 'GraphRAG', '사고 인과 그래프 (중대재해 인과 추적)', '건설', '★★'),
    (45, 'Memory Architectures', 'AI 장기기억 (직원-고객 행동 누적)', '부서', '★★'),
    (46, 'Memory Architectures', '위험 사례 영구 학습', '부서', '★★'),
    (47, 'Prompt Injection', '직원 LLM 사용 보호 (데이터유출 방어)', '부서', '★★★'),
    (48, 'Prompt Injection', '고객 챗봇 방어', '제품', '★★'),
    (49, 'Mech. Interpretability', '위험점수 설명가능성 (EU AI Act 의무)', '부서', '★★★'),
    (50, 'Mech. Interpretability', '규제 보고 AI 투명성', 'GRC', '★★'),
    (51, 'Deepfake Detection', '보이스피싱 방어 (2026 의무)', '금융 VMS', '★★★'),
    (52, 'C2PA', '회계-법무 문서 출처 검증', '내부회계', '★★'),
    (53, 'Federated Learning', '여러 고객사 데이터 안 모으고 학습', '부서', '★★★'),
    (54, 'Federated Learning', '금융권 컨소시엄 부정탐지', '금융', '★★'),
    (55, 'Confidential Compute', '고객 데이터 분석 시 격리', '부서', '★★'),
    (56, 'Confidential Compute', 'AI 모델 격리 실행', '보안', '★★'),
    (57, 'Synthetic Data', '행동AI 학습데이터 무한 생성', '부서', '★★★'),
    (58, 'Synthetic Data', '개인정보 없이 모델 학습', '데이터', '★★'),
    (59, 'Affective Computing', '카지노 딜러 스트레스 사전 감지', '카지노 VMS', '★★★'),
    (60, 'Affective Computing', '금융직원 부정 사전 예방', '부서-내부회계', '★★★'),
    (61, 'Adversarial ML', 'AI 모델 강건성 테스트 서비스', '부서', '★★'),
    (62, 'Adversarial ML', '부서 AI 모델 자체 보호', '부서', '★★'),
    (63, 'AI Workload Protection', '한국 1호 MSP 운영 (Palo Alto OEM)', '그룹 시너지', '★★★'),
    (64, 'AI Workload Protection', '내부 AI 보안 표준', '전 제품', '★★'),
    (65, 'NHI', 'AI 에이전트-서비스계정 신원관리', '보안', '★★'),
    (66, 'NHI', '서비스계정 자동 만료-교체', '보안', '★'),
    (67, 'CSMA', 'ITCEN CORE 제품군 분산 보안 표준', '전 제품', '★★'),
    (68, 'CSMA', '멀티클라우드 보안 통합 (CTS 협업)', 'CTS 협업', '★★'),
    (69, 'DSPM', '고객사 데이터 위치-민감도 추적', '데이터', '★★'),
    (70, 'DSPM', '개인정보 자동 분류-라벨링', '데이터', '★★'),
    (71, 'Behavioral Biometrics', '부서 행동분석의 자연 확장', '부서 IP', '★★★'),
    (72, 'Behavioral Biometrics', 'VMS 상시 인증', 'VMS', '★★'),
    (73, 'Continuous Auth', '금융권 차세대 인증', '금융', '★★'),
    (74, 'Continuous Auth', '내부자위협 방어 (UEBA + 상시인증)', '부서', '★★★'),
    (75, 'Passkeys', 'ITCEN CORE 제품군 차세대 인증', '전 제품', '★★'),
    (76, 'Passkeys', '카지노-금융 비밀번호 완전 제거', 'VMS', '★★'),
    (77, 'DPO', '부서 행동위험 모델 효율적 학습', '부서', '★★'),
    (78, 'DPO', '산업별 sLLM 저비용 fine-tuning', '전 제품', '★★'),
    (79, 'RLHF', '도메인 전문가 피드백 모델 정교화', '부서', '★★'),
    (80, 'RLHF', '부서 위험판단 모델 최적화', '부서', '★★'),
    (81, 'Constitutional AI', '부서 SOP-법규를 헌법으로 박기', '부서', '★★★'),
    (82, 'Constitutional AI', 'ITCEN CORE 제품 컴플라이언스 자동', '전 제품', '★★'),
    (83, 'LoRA', '각 산업-고객 맞춤 sLLM 저비용', '전 제품', '★★'),
    (84, 'LoRA', '부서 모델 빠른 업데이트', '부서', '★★'),
    (85, 'Vector DB', '행동패턴 임베딩 검색', '부서', '★★'),
    (86, 'Vector DB', '회계 부정 패턴 검색', '내부회계', '★★'),
    (87, 'HyDE', '위험 질의 답 먼저 만들고 검색', '부서', '★★'),
    (88, 'HyDE', '법규 적합성 자동 추론', 'GRC', '★★'),
    (89, 'Long Context (1M+)', '1년치 행동 데이터 통합 분석 (단일 호출)', '부서', '★★'),
    (90, 'Long Context', '대형 계약서-법규 통합 검토', 'GRC', '★★'),
    (91, 'AI Governance', 'ITCEN CORE 모든 AI 통합 관리', '전 제품', '★★'),
    (92, 'AI Governance', 'ISO 42001 인증 자동화', 'GRC', '★★★'),
    (93, 'Bias Detection', '부서 행동점수 편향 검출', '부서', '★★'),
    (94, 'Explainability', '위험판단 설명 가능 (EU AI Act 의무)', '부서', '★★★'),
    (95, 'Domain FM', '회계 특화 LLM (BloombergGPT 한국판)', '내부회계', '★★★'),
    (96, 'Domain FM', '카지노 도메인 LLM', '카지노', '★★'),
    (97, 'AI Search', '사내 위험정보 자율 검색', '부서', '★★'),
    (98, 'AI Search', '컴플라이언스 자율 조사', 'GRC', '★★'),
    (99, 'Ambient Intelligence', '저비용 IoT + 디지털트윈 통합', '디지털트윈', '★★'),
    (100, 'Ambient Intelligence', '산업 안전 자동 모니터링 (중대재해법)', '건설', '★★★'),
]

for i, (n, tech, idea, asset, syn) in enumerate(APPS_100, start=4):
    ws6.cell(row=i, column=1, value=n).border = brd
    c = ws6.cell(row=i, column=2, value=tech)
    c.font = Font(name='맑은 고딕', size=9, bold=True); c.border = brd
    c = ws6.cell(row=i, column=3, value=idea)
    c.font = cell_font; c.alignment = Alignment(wrap_text=True, vertical='center'); c.border = brd
    c = ws6.cell(row=i, column=4, value=asset)
    c.font = cell_font; c.alignment = Alignment(horizontal='center', vertical='center'); c.border = brd
    c = ws6.cell(row=i, column=5, value=syn)
    c.alignment = Alignment(horizontal='center', vertical='center'); c.border = brd
    ws6.row_dimensions[i].height = 22
print(f'Sheet 6 (100 apps): 100 rows')

# Sheet 7: AI Risk Lighthouse
if '7.AIRiskLighthouse' in wb.sheetnames:
    del wb['7.AIRiskLighthouse']
ws7 = wb.create_sheet('7.AIRiskLighthouse')
ws7['A1'] = 'AI Risk Lighthouse — 부서 핵심 IP 신사업 (한국 표준화 전략)'
ws7['A1'].font = Font(name='맑은 고딕', size=14, bold=True, color='C00000')
ws7.merge_cells('A1:D1')
ws7['A2'] = 'Google Lighthouse 가 웹페이지 점수 매기듯, 회사의 AI-내부통제-행동위험을 자동 감사-점수 매김. ITCEN CORE 가 한국 표준 만들면 모든 한국 기업 의무 도입 잠재.'
ws7['A2'].font = Font(name='맑은 고딕', size=10, italic=True, color='595959')
ws7['A2'].alignment = Alignment(wrap_text=True)
ws7.merge_cells('A2:D2')
ws7.row_dimensions[2].height = 40

ws7.column_dimensions['A'].width = 6
ws7.column_dimensions['B'].width = 22
ws7.column_dimensions['C'].width = 60
ws7.column_dimensions['D'].width = 18

hdrs7 = ['#', '카테고리', '검사 항목', '가중치']
for col, h in enumerate(hdrs7, 1):
    c = ws7.cell(row=4, column=col, value=h)
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', vertical='center'); c.border = brd

LIGHTHOUSE = [
    (1, 'Self-Critique', 'AI 결정에 2단계 검증-confidence 점수 공개-재시도 루프', '15%'),
    (2, 'Causal AI', '위험점수에 인과 설명-단순 상관관계 X-DoWhy 적용', '15%'),
    (3, 'Behavioral Coverage', 'UEBA-VMS-CCTV-결재-감정 데이터 통합', '15%'),
    (4, 'Interpretability', 'AI 결정 설명 가능-EU AI Act 의무-SHAP-LIME', '12%'),
    (5, 'Privacy (PET)', 'Federated-Confidential-동형암호-차등프라이버시', '12%'),
    (6, 'Compliance', '한국 법규-EU AI Act-금감원 가이드 자동 추적', '10%'),
    (7, 'Quality (FP/FN)', '거짓양성-거짓음성 자동 모니터-정기 audit', '11%'),
    (8, 'Self-Improvement', '실패 사례 학습-반복 방지-Reflexion 루프', '10%'),
]

for i, (n, cat, item, w) in enumerate(LIGHTHOUSE, start=5):
    ws7.cell(row=i, column=1, value=n).border = brd
    c = ws7.cell(row=i, column=2, value=cat)
    c.font = Font(name='맑은 고딕', size=10, bold=True, color='C00000'); c.border = brd
    c = ws7.cell(row=i, column=3, value=item)
    c.font = cell_font; c.alignment = Alignment(wrap_text=True, vertical='center'); c.border = brd
    c = ws7.cell(row=i, column=4, value=w)
    c.alignment = Alignment(horizontal='center', vertical='center'); c.border = brd
    ws7.row_dimensions[i].height = 28

ws7.cell(row=14, column=1, value='3년 사업화 로드맵').font = Font(name='맑은 고딕', size=12, bold=True, color='4472C4')
ws7.merge_cells('A14:D14')

ROADMAP = [
    ('2026 Q4', 'IP 확보', '부서가 Lighthouse 8 카테고리 + 점수 모델 IP 확보. SOP 헌법화'),
    ('2027 Q1', '베타 출시', 'ITCEN CORE 내부회계-GRC 고객 베타 (100社). 무료 진단 -> 유료'),
    ('2027 Q2', '표준 lobby', 'KISA-금감원-개인정보위 협력. K-AI Risk Standard 초안'),
    ('2027 Q3', '공식 표준', '한국 K-AI Risk Lighthouse 공식 표준 등록. 인증 사업'),
    ('2028+', '의무 도입', '모든 한국 대기업 의무 (5000社 × 1억 = 5천억 영구)'),
    ('2029+', '글로벌', 'K-Standard -> 동남아-중동-EU 수출. 글로벌 라이선싱'),
]

for i, (q, phase, detail) in enumerate(ROADMAP, start=15):
    c = ws7.cell(row=i, column=1, value=q)
    c.font = Font(name='맑은 고딕', size=10, bold=True, color='4472C4'); c.border = brd
    c = ws7.cell(row=i, column=2, value=phase)
    c.font = Font(name='맑은 고딕', size=10, bold=True); c.border = brd
    c = ws7.cell(row=i, column=3, value=detail)
    c.font = cell_font; c.alignment = Alignment(wrap_text=True, vertical='center'); c.border = brd
    ws7.cell(row=i, column=4, value='').border = brd
    ws7.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
    ws7.row_dimensions[i].height = 32

print(f'Sheet 7: {len(LIGHTHOUSE)} cats + {len(ROADMAP)} roadmap')

wb.save(src)
print(f'\nSAVED: {src}')
print(f'Total sheets: {wb.sheetnames}')
