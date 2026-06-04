"""ITcen 공모전 제출용 Excel 생성.

4 시트:
  1. ROI 매트릭스 (분류·top 추천·근거)
  2. 100 아이디어 (카테고리·번호·제목·설명·ROI 분류)
  3. 제출 양식 (빈칸 템플릿)
  4. BMC 예시 — 가장 유망한 #2 sLLM 회의록·법무문서 자동 요약·태깅
"""
from __future__ import annotations
import sys
import subprocess
from pathlib import Path
from datetime import date

# openpyxl 자동 install
try:
    import openpyxl  # noqa: F401
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"], check=True)
    import openpyxl  # noqa: F401

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
OUT_DIR = PROJECT_ROOT / "outputs" / "itcen"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_FILE = OUT_DIR / f"itcen-proposal-{date.today().isoformat()}.xlsx"

# ───────── 스타일 ─────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
CAT_FILL = PatternFill("solid", fgColor="D9E2F3")
CAT_FONT = Font(name="맑은 고딕", bold=True, size=11, color="1F4E79")
NORMAL = Font(name="맑은 고딕", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = Workbook()

# ═══════════════════════════════════════════════════════════
# Sheet 1: ROI 매트릭스
# ═══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "1.ROI매트릭스"

ROI_MATRIX = [
    ("즉시 매출 (6~12개월)",
     "#1 금융회의록 AI / #2 공공정책문서 / #3 RAG포털 / #6 규제추적 / #41 FDS이상거래 / #42 AML보고 / #91 사내RAG / #92 AI회의록",
     "ITCEN 기존 금융·공공SI 고객 직접판매. 레퍼런스 활용. 객단가 5~12억", "★★★★★"),
    ("단기 PoC → 1년 매출",
     "#11 ZeroTrust정책 / #12 보안이벤트 / #22 설비예지 / #43 신용평가 / #54 화재감지 / #93 프로젝트자동화",
     "ITCEN 보안(PNS)·클라우드(CTS)·공공(ENTEC) 사업부 협력. 금융감시 정기발주", "★★★★"),
    ("중장기 R&D (정부매칭)",
     "#71 PQC양자암호 / #72 QKD백본 / #73 FHE동형암호 / #75 CBDC결제 / #86 UAM관제",
     "과기정통부·금감위·산림청 과제 매칭. K-보안·K-핀테크 국가전략과제", "★★★"),
    ("B2G 공공입찰 강점",
     "#31 스마트시티관제 / #32 CCTV감지 / #33 환경센서 / #35 화재감시 / #39 농업IoT / #55 지진시뮬레이션 / #68 산림위성",
     "행정안전부·환경부·산림청·농림부 발주. ITCEN ENTEC(전자정부·NEIS) 채널 활용", "★★★★★"),
    ("글로벌 수출 가능",
     "#71 PQC양자암호 / #72 QKD / #82 V2X보안 / #84 라스트마일로봇 / #88 자율주행익명화",
     "미·EU 규제 대응. K-방산·자동차·금융 국제기준. ITCEN 글로벌법인 진출 플랫폼", "★★★★"),
    ("레드오션 — 신규투자 제한",
     "범용AI챗봇·일반RPA·일반데이터품질·주차장예측 등",
     "이미 마이크로소프트·구글·삼성 등 점유. ITCEN은 도메인특화에 집중", "⚠"),
    ("카테고리별 강점시너지 (ITCEN 계열사 활용)",
     "금융IT(#1·#41·#42·#43·#44·#48) / 보안(#11·#12·#71·#72) / 공공(#31·#32·#54·#55·#59·#68) / Web3(#74·#76·#80) / 모빌리티(#82·#84·#86·#88)",
     "같은 컨소시엄으로 묶어 대형 정부프로젝트·B2G 대규모 수주 가능. ITCEN 계열사 간 협력의뢰시스템 구축", "★★★★★"),
]

ws["A1"] = "ITcen 공모전 — 100 아이디어 ROI 매트릭스"
ws["A1"].font = Font(name="맑은 고딕", bold=True, size=14, color="1F4E79")
ws.merge_cells("A1:D1")

headers = ["분류", "Top 추천", "근거", "추천 강도"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

for i, (cls, picks, reason, star) in enumerate(ROI_MATRIX, 4):
    ws.cell(row=i, column=1, value=cls).font = CAT_FONT
    ws.cell(row=i, column=1).fill = CAT_FILL
    ws.cell(row=i, column=2, value=picks).font = NORMAL
    ws.cell(row=i, column=3, value=reason).font = NORMAL
    ws.cell(row=i, column=4, value=star).font = NORMAL
    for col in range(1, 5):
        ws.cell(row=i, column=col).alignment = WRAP
        ws.cell(row=i, column=col).border = BORDER

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 40
ws.column_dimensions["D"].width = 12
for i in range(4, 4 + len(ROI_MATRIX)):
    ws.row_dimensions[i].height = 50

# ═══════════════════════════════════════════════════════════
# Sheet 2: 100 아이디어
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2.100아이디어")

IDEAS = [
    # (번호, 카테고리, 제목, ROI 분류, 예상 객단가, 추천)
    ("A. AI·LLM 응용 (ITCEN SI 활용)", [
        (1, "금융권 회의록·합의서 AI 자동 요약·태깅 (ITCEN 금융SI 레퍼런스)", "즉시 매출", "5~10억", "★★★★★"),
        (2, "공공기관 정책문서·법률검토 AI 분석 (ITCEN 공공SI 확장)", "즉시 매출", "3~8억", "★★★★★"),
        (3, "엔터프라이즈 RAG 포털 (기존 SI 고객·금융·공공 내부화)", "즉시 매출", "5~12억", "★★★★★"),
        (4, "보안 인시던트 자동 분류·원인분석 (ITCEN PNS 보안팀 협업)", "단기 PoC", "5~12억", "★★★★"),
        (5, "클라우드 인프라 AI 최적화 (비용·성능·보안) (ITCEN CTS 클라우드강점)", "단기 PoC", "3~7억", "★★★★"),
        (6, "금융감시 규제변경 자동 추적·내부영향 분석 (정규직무)", "즉시 매출", "3~7억", "★★★★"),
        (7, "산업제어망(OT) AI 이상탐지 (ITCEN 공공·제조 SI 대상)", "단기 PoC", "8~15억", "★★★★"),
        (8, "디지털 정부 업무 자동화 (행정안전부·NEIS 강점)", "B2G", "5~12억", "★★★★"),
        (9, "SI 프로젝트 우선순위 AI 자동 배분 (ITCEN 자사 효율화)", "단기 PoC", "1~3억", "★★★"),
        (10, "법무사례·판례 AI 검색 (법무법인·기업법무 협력)", "단기 PoC", "2~5억", "★★★"),
    ]),
    ("B. 사이버보안·ITCEN PNS 중심", [
        (11, "Zero Trust 정책 AI 자동생성 (금융권 격리인증) (ITCEN PNS)", "즉시 매출", "10~20억", "★★★★★"),
        (12, "금융감시 보안이벤트 자동 분류·위험도 점수 (ISMS-P)", "즉시 매출", "5~10억", "★★★★★"),
        (13, "클라우드 보안 설정 AI 감시·자동 정정 (ITCEN CTS)", "단기 PoC", "3~8억", "★★★★"),
        (14, "공공기관 보안 취약점 스캔 자동화 (정부보안기준)", "B2G", "5~12억", "★★★★"),
        (15, "데이터 마스킹 정책 AI 자동생성 (개인정보보호법)", "단기 PoC", "3~7억", "★★★"),
        (16, "침해사고 근인분석 AI 보고서 생성 (SOAR + LLM)", "단기 PoC", "5~12억", "★★★★"),
        (17, "API 보안 자동 검증 (OAuth/OIDC 형태 분석)", "단기 PoC", "2~5억", "★★★"),
        (18, "산업제어망 SCADA 보안 모니터링 (제조·에너지)", "B2G", "8~15억", "★★★★"),
        (19, "사이버위험 보험 청구 자동화 (손해사정)", "단기 PoC", "3~7억", "★★★"),
        (20, "금융권 컴플라이언스 AI 트레이닝 (정기 의무교육)", "단기 PoC", "1~3억", "★★"),
    ]),
    ("C. 데이터·분석·ITCEN CTS클라우드", [
        (21, "금융 고객 신용예측 AI (대안데이터 활용)", "즉시 매출", "5~12억", "★★★★★"),
        (22, "제조·에너지 설비 예지보전 (IoT 센서 + 시계열AI)", "단기 PoC", "8~15억", "★★★★"),
        (23, "고객탈출 예측·재구매 캠페인 자동화 (금융권)", "단기 PoC", "3~7억", "★★★★"),
        (24, "공공데이터 통합 분석 대시보드 (행정·사회통계)", "B2G", "3~8억", "★★★★"),
        (25, "클라우드 비용 AI 최적화 (AWS·Azure 자동 우측규모)", "단기 PoC", "2~5억", "★★★"),
        (26, "데이터 카탈로그 + 자동 품질점수 (메타데이터)", "단기 PoC", "2~5억", "★★★"),
        (27, "합성데이터 생성 (개인정보 익명화·규제샘플)", "단기 PoC", "3~7억", "★★★"),
        (28, "규제보고 자동화 (금감위·금융감시 정기 리포트)", "즉시 매출", "5~12억", "★★★★"),
        (29, "매출·손익 이상 자동진단 (AI 근인분석)", "단기 PoC", "2~5억", "★★★"),
        (30, "환율·원자재 가격 예측 (수출입기업/무역금융용)", "단기 PoC", "2~5억", "★★★"),
    ]),
    ("D. IoT·스마트시티·ITCEN ENTEC공공", [
        (31, "스마트시티 통합 관제센터 (CCTV·센서·신고 통합)", "B2G", "15~40억", "★★★★★"),
        (32, "도시 CCTV 밀집도·이상행동 자동 감지 (공안·재난)", "B2G", "8~20억", "★★★★★"),
        (33, "공공 건물 에너지 효율화 AI (학교·청사·도서관)", "B2G", "5~15억", "★★★★"),
        (34, "환경센서 실시간 대기질·하천수질 모니터링", "B2G", "5~12억", "★★★★"),
        (35, "화재감시 영상 AI (산불·건물화재 조기탐지)", "B2G", "8~15억", "★★★★★"),
        (36, "제조업 산업안전 PPE 미착용 탐지 (엣지AI)", "단기 PoC", "3~8억", "★★★★"),
        (37, "지하철·버스 혼잡도 예측·안내 (대중교통)", "B2G", "5~12억", "★★★★"),
        (38, "도시 물류 최적경로 (배송·쓰레기·수도 통합)", "B2G", "5~12억", "★★★"),
        (39, "농업 IoT (병해충 탐지·관개 자동) (지자체)", "B2G", "3~8억", "★★★"),
        (40, "공항·항만 자동화 (화물추적·적재최적화)", "B2G", "10~25억", "★★★★"),
    ]),
    ("E. 금융·규제·ITCEN 금융IT강점", [
        (41, "AI 이상거래 탐지 FDS (금융감시 기준) (ITCEN SI레퍼런스)", "즉시 매출", "8~15억", "★★★★★"),
        (42, "자금세탁방지 AML 자동 의심보고 (금감위 정기)", "즉시 매출", "8~15억", "★★★★★"),
        (43, "신용평가 대체데이터 (소셜·통신·결제)", "단기 PoC", "5~12억", "★★★★"),
        (44, "금융감시 규제변경 자동 추적·영향분석 (정규)", "즉시 매출", "5~10억", "★★★★"),
        (45, "보험사기 탐지 AI + 손해사정 보조 (보험사)", "단기 PoC", "5~12億", "★★★★"),
        (46, "마이데이터 자산통합 (금융권 협력)", "단기 PoC", "3~7억", "★★★"),
        (47, "결제사기 예방 AI (카드·계좌·간편결제)", "단기 PoC", "3~7억", "★★★"),
        (48, "금리·환율 리스크 자동 보고서", "단기 PoC", "2~5억", "★★★"),
        (49, "금융상품 가입자 적합성 AI 검증", "단기 PoC", "2~5억", "★★★"),
        (50, "차세대 코어뱅킹 마이그레이션 자동화", "B2G", "10~25억", "★★★★"),
    ]),
    ("F. 공공·보건·재난·ITCEN ENTEC", [
        (51, "의료영상 AI 판독보조 (병원·보건소)", "단기 PoC", "5~12억", "★★★★"),
        (52, "응급실 환자분류 AI (triage) (보건청)", "B2G", "3~8억", "★★★★"),
        (53, "산업안전 낙상·충돌 실시간 감지 (작업장)", "단기 PoC", "3~8억", "★★★★"),
        (54, "화재·연기 CCTV 조기탐지 (소방청·건물)", "B2G", "8~20억", "★★★★★"),
        (55, "지진·홍수 시뮬레이션 (재난청)", "B2G", "5~15억", "★★★★"),
        (56, "화학물질 누출 시뮬레이션 (환경부·산단)", "B2G", "3~8억", "★★★"),
        (57, "만성질환 원격모니터링 (건강보험공단)", "B2G", "5~12억", "★★★★"),
        (58, "정신건강 위기상담 AI (보건복지부)", "B2G", "2~5억", "★★★"),
        (59, "재난 SNS 분석·피해지도 (행안부)", "B2G", "5~12억", "★★★★"),
        (60, "공공의료 인력 최적배치 AI (지역균형)", "B2G", "3~8억", "★★★"),
    ]),
    ("G. ESG·기후·에너지·공공/기업경영", [
        (61, "Scope 1·2·3 탄소배출 자동산정·보고서 (정규의무)", "즉시 매출", "3~8억", "★★★★"),
        (62, "ESG 공급망 평가 AI (협력사 스코어링)", "단기 PoC", "3~7억", "★★★★"),
        (63, "재생에너지 RE100 추적·거래·인증", "단기 PoC", "5~12억", "★★★"),
        (64, "전력수요 예측·수요반응 자동참여 (산단)", "단기 PoC", "3~7억", "★★★"),
        (65, "폐기물 분류 영상AI (도시·산단)", "B2G", "2~5억", "★★★"),
        (66, "그린빌딩 에너지 자동모니터링 (공공청사)", "B2G", "3~8억", "★★★"),
        (67, "산림보호 위성 모니터링 (산림청)", "B2G", "5~12억", "★★★★"),
        (68, "수질·수량 IoT 모니터링 (환경부·지자체)", "B2G", "5~12억", "★★★★"),
        (69, "마이크로그리드 자동운영 (지역기반)", "중장기 R&D", "5~12억", "★★★"),
        (70, "탄소배출권 거래 플랫폼 (K-ETS확대)", "B2G", "10~25억", "★★★★"),
    ]),
    ("H. 양자암호·차세대보안·Web3", [
        (71, "양자내성암호 PQC 전환도구 (CRYPTO인벤토리)", "글로벌 수출", "8~15억", "★★★★★"),
        (72, "양자키분배 QKD 금융권백본 (보안등급)", "중장기 R&D", "15~40억", "★★★★"),
        (73, "동형암호 FHE 의료/금융 데이터분석", "중장기 R&D", "8~15억", "★★★★"),
        (74, "블록체인 기반 DID 신원증명 (부산거래소)", "단기 PoC", "5~12억", "★★★★"),
        (75, "CBDC 결제 인프라 (한국은행·금감위)", "중장기 R&D", "15~40억", "★★★★"),
        (76, "토큰증권 거래플랫폼 (ITCEN 기업공개전략)", "단기 PoC", "10~25억", "★★★★"),
        (77, "영지식증명 zk-SNARK 금융용 본인인증", "중장기 R&D", "5~12억", "★★★"),
        (78, "컨피덴셜컴퓨팅 TEE 클라우드분석 (금융)", "단기 PoC", "8~15억", "★★★"),
        (79, "Web3 스마트계약 감시·위험탐지", "단기 PoC", "3~8억", "★★★"),
        (80, "암호화폐·NFT 거래 컴플라이언스AI", "단기 PoC", "5~12억", "★★★★"),
    ]),
    ("I. 자율주행·모빌리티·교통", [
        (81, "자율주행 시뮬레이션 시나리오 자동생성 (LLM)", "단기 PoC", "5~12억", "★★★★"),
        (82, "V2X 보안 (차-인프라 인증서 자동발급)", "글로벌 수출", "8~15억", "★★★★"),
        (83, "화물트럭 군집주행 Platooning (산단·고속도)", "중장기 R&D", "15~40억", "★★★"),
        (84, "라스트마일배송 로봇 운영 플랫폼 (배송사)", "단기 PoC", "8~15억", "★★★★"),
        (85, "드론배송 항로·공역관리 (우정사업본부)", "B2G", "8~15억", "★★★★"),
        (86, "도시항공교통 UAM 관제시뮬레이션", "중장기 R&D", "15~40억", "★★★"),
        (87, "전기차충전 동적가격·예약 (에너지회사)", "단기 PoC", "3~8억", "★★★"),
        (88, "자율주행 사고데이터 익명화·공유표준 (K-방산)", "글로벌 수출", "5~12억", "★★★★"),
        (89, "운전자 졸음·산만감지 (버스·택시)", "단기 PoC", "3~8억", "★★★"),
        (90, "공유모빌리티 수요예측·최적배치 (시내교통)", "단기 PoC", "3~7억", "★★★"),
    ]),
    ("J. 디지털워크·자동화·ITCEN 효율화", [
        (91, "사내 RAG 컨시어지 (Slack/Teams/Jira통합)", "즉시 매출", "3~8억", "★★★★★"),
        (92, "AI 회의록+액션자동할당 (정기·프로젝트회의)", "즉시 매출", "2~5억", "★★★★★"),
        (93, "SI프로젝트 자동화 플로우 빌더 (ITCEN자사용)", "단기 PoC", "3~7억", "★★★★"),
        (94, "임직원 질의응답 멀티에이전트 (HR·IT·총무통합)", "단기 PoC", "2~5억", "★★★★"),
        (95, "시스템 아키텍처 코드→다이어그램 자동생성", "단기 PoC", "1~3억", "★★★"),
    ]),
    ("K. 리스크모니터링·행동위험분석 (ITCEN CORE 직결 ⭐핵심부서⭐)", [
        (96, "양자센싱 기반 물리침입 탐지 (양자자기장·진동·전자기) — 데이터센터·금융전산실·공공기관 출입통제 고도화", "즉시 매출", "10~25억", "★★★★★"),
        (97, "피지컬센서 + 행동AI 통합관제 (CCTV+IoT+생체+동선) — 금융권 내부자위협·물리보안 ZeroTrust", "즉시 매출", "15~30억", "★★★★★"),
        (98, "직원 행동위험분석 (UEBA — User & Entity Behavior Analytics) — 부정거래·정보유출·이상로그인 AI 상시감시", "즉시 매출", "8~20억", "★★★★★"),
        (99, "양자내성암호 PQC + 행동AI 융합 — Post-Quantum 시대 ITCEN CORE 리스크 플랫폼 (금융감독원 차세대 표준)", "글로벌 수출", "20~50억", "★★★★★"),
        (100, "Insider Threat Hunting AI — 행동패턴 클러스터링·이상점수·자동조사 워크플로우 (ITCEN CORE 리스크모니터링부 핵심 IP)", "즉시 매출", "12~25억", "★★★★★"),
    ]),
]

ws2["A1"] = "ITcen 공모전 — 100 아이디어 전체 리스트"
ws2["A1"].font = Font(name="맑은 고딕", bold=True, size=14, color="1F4E79")
ws2.merge_cells("A1:F1")

headers2 = ["번호", "카테고리", "제목·핵심", "ROI 분류", "예상 객단가", "추천 강도"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER

row = 4
for cat, items in IDEAS:
    # 카테고리 row
    cc = ws2.cell(row=row, column=1, value=cat)
    cc.fill = CAT_FILL
    cc.font = CAT_FONT
    cc.alignment = Alignment(horizontal="left", vertical="center")
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cc.border = BORDER
    row += 1
    for num, title, roi, price, star in items:
        ws2.cell(row=row, column=1, value=num).font = NORMAL
        ws2.cell(row=row, column=2, value=cat[:1]).font = NORMAL
        ws2.cell(row=row, column=3, value=title).font = NORMAL
        ws2.cell(row=row, column=4, value=roi).font = NORMAL
        ws2.cell(row=row, column=5, value=price).font = NORMAL
        ws2.cell(row=row, column=6, value=star).font = NORMAL
        for col in range(1, 7):
            ws2.cell(row=row, column=col).alignment = WRAP
            ws2.cell(row=row, column=col).border = BORDER
        row += 1

ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 8
ws2.column_dimensions["C"].width = 60
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 12
ws2.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════
# Sheet 3: 제출 양식 (빈칸)
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3.제출양식")

ws3["A1"] = "ITcen 공모전 — 참가 신청 양식"
ws3["A1"].font = Font(name="맑은 고딕", bold=True, size=14, color="1F4E79")
ws3.merge_cells("A1:B1")

FORM_FIELDS = [
    ("팀명", "*필수", ""),
    ("소속 부서", "*필수", "예: 기술혁신본부"),
    ("팀원 수", "*필수", "예: 3명"),
    ("사업 분야", "*필수", "예: AI 기반 고객서비스 자동화, 스마트팩토리 플랫폼 등"),
    ("팀원 이름", "*필수", "예: 홍길동, 김철수, 이영희"),
    ("연락처 이메일", "*필수", "team@itcen.com"),
    ("프로젝트 제목", "*필수", "아이디어의 핵심을 담은 제목"),
    ("아이디어 설명", "*필수", "해결하려는 문제, 제안하는 솔루션, 핵심 기능"),
    ("기대 효과", "선택", "사업적 효과·성과"),
    ("아이디어 제안서 URL", "선택", "https://docs.google.com/... 또는 https://github.com/..."),
    ("아이디어 제안서 파일", "선택", "PPT·PPTX·PDF (최대 10MB)"),
    ("목업 URL", "선택", "https://www.figma.com/... 또는 데모 사이트"),
    ("BMC 문서 URL", "선택", "Google Slides·Notion·Figma 공유 링크"),
    ("BMC 파일", "선택", "PDF·PPT·PPTX·PNG·JPG (최대 10MB)"),
    ("수정용 4자리 PIN", "*필수", "추후 제출 결과물 수정 시 사용"),
]

ws3.cell(row=3, column=1, value="항목").fill = HEADER_FILL
ws3.cell(row=3, column=1).font = HEADER_FONT
ws3.cell(row=3, column=2, value="필수 여부").fill = HEADER_FILL
ws3.cell(row=3, column=2).font = HEADER_FONT
ws3.cell(row=3, column=3, value="입력 / 안내").fill = HEADER_FILL
ws3.cell(row=3, column=3).font = HEADER_FONT
for col in range(1, 4):
    ws3.cell(row=3, column=col).alignment = CENTER
    ws3.cell(row=3, column=col).border = BORDER

for i, (label, req, hint) in enumerate(FORM_FIELDS, 4):
    ws3.cell(row=i, column=1, value=label).font = CAT_FONT
    ws3.cell(row=i, column=1).fill = CAT_FILL
    ws3.cell(row=i, column=2, value=req).font = NORMAL
    ws3.cell(row=i, column=3, value=hint).font = NORMAL
    for col in range(1, 4):
        ws3.cell(row=i, column=col).alignment = WRAP
        ws3.cell(row=i, column=col).border = BORDER

ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 60

# ═══════════════════════════════════════════════════════════
# Sheet 4: BMC 예시 (#1 금융권 회의록·합의서 AI 자동 요약·태깅)
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4.BMC예시-1번")

ws4["A1"] = "BMC 예시 — #1 금융권 회의록·합의서 AI 자동 요약·태깅 (ITCEN 금융SI 레퍼런스활용)"
ws4["A1"].font = Font(name="맑은 고딕", bold=True, size=14, color="1F4E79")
ws4.merge_cells("A1:B1")

ws4["A2"] = "ROI: 즉시 매출 (6~12개월) · 객단가 5~10억 · 추천 ★★★★★ · ITCEN SI 기존고객 대상"
ws4["A2"].font = Font(name="맑은 고딕", italic=True, size=10, color="555555")
ws4.merge_cells("A2:B2")

BMC = [
    ("01 핵심 파트너 (Key Partners)",
     "• Anthropic / OpenAI 금융전용 기금모델 라이선스\n"
     "• 국내금융감시위원회·은행권·증권협회 표준자문\n"
     "• Naver HyperCLOVA X (한국어 금융도메인)\n"
     "• ITCEN 금융SI팀 (ENTEC·CTS 기존 고객망)\n"
     "• 대형로펌·금융감시 컨설턴트 (학습데이터+검증)"),
    ("02 핵심 활동 (Key Activities)",
     "• 금융권 회의록·합의서·감시보고 데이터 수집(익명화)\n"
     "• sLLM fine-tune (금융용 8B~13B모델 + 한국어)\n"
     "• 규제·컴플라이언스 태깅 규칙 자동화\n"
     "• 정확도·환각률 정기평가 (금감위 기준)\n"
     "• ITCEN SI통합 (금융ERP·시스템+MS Teams/Slack)\n"
     "• 정보보안인증 (ISMS-P, K-FSI, PCI-DSS)"),
    ("03 핵심 자원 (Key Resources)",
     "• 금융권 회의록·계약·판결데이터 (1M+)\n"
     "• GPU 추론인프라 (L4/A10, 낮은지연)\n"
     "• ITCEN 금융SI인력 (300+명)\n"
     "• 기존 금융고객 (은행·증권·보험·금감위)\n"
     "• 보안인증·규제대응 전담팀"),
    ("04 가치 제안 (Value Propositions)",
     "• 감시보고 작성시간 70% 절감 (일일자동작성)\n"
     "• 회의내용 즉시 액션/이슈 추출 (누락방지)\n"
     "• 금융감시 규정 자동 준수 검증\n"
     "• VPC격리·On-prem가능 (민감정보보호)\n"
     "• 3개월 PoC후 정규배포 (금감위 사후평가)"),
    ("05 고객 관계 (Customer Relationships)",
     "• ITCEN 기존금융SI 고객 직접영업 (CRO/컴플라이언스)\n"
     "• 분기별 정기리뷰·SLA관리\n"
     "• 고객피드백 기반 모델재학습\n"
     "• 금융감시 규제변경 자동반영\n"
     "• 연1~2회 금융IT컨퍼런스 지원"),
    ("06 채널 (Channels)",
     "• ITCEN ENTEC/CTS 금융SI팀 직영업\n"
     "• 나라장터·금감위 공공조달\n"
     "• 은행권·증권협회 추천\n"
     "• 금융IT 컨퍼런스·세미나\n"
     "• AWS·Azure금융마켓플레이스"),
    ("07 고객 세그먼트 (Customer Segments)",
     "• 1차: 메가뱅크·증권사 (감시보고·규제대응)\n"
     "• 2차: 보험사·카드사 (거래감시·AML)\n"
     "• 3차: 금감위·중앙은행 (규제감시)\n"
     "• 4차: 대형로펌·금융법무팀\n"
     "• 5차: 금융감시솔루션 납품사"),
    ("08 비용 구조 (Cost Structure)",
     "• Foundation 모델 API (월 2억)\n"
     "• GPU 추론 (월 5천만)\n"
     "• ML엔지니어 (8명×1.2억 = 9.6억/년)\n"
     "• 정보보안·규제대응 (월 1천만)\n"
     "• ITCEN SI통합·지원 (월 5천만)\n"
     "• 영업·마케팅 (연매출 12%)"),
    ("09 수익 흐름 (Revenue Streams)",
     "• 연간 SaaS 라이선스 (사당 3~5억)\n"
     "• SI 통합·커스터마이징 (사당 2~5억)\n"
     "• 도메인 fine-tune 추가비용 (사당 1~3억)\n"
     "• On-premise라이선스 (3년 5~12억)\n"
     "• 관리형서비스 (월 5천만~1억)\n"
     "• 컨설팅·교육·지원 (시간당 300만)"),
]

ws4.cell(row=4, column=1, value="9 항목").fill = HEADER_FILL
ws4.cell(row=4, column=1).font = HEADER_FONT
ws4.cell(row=4, column=2, value="내용").fill = HEADER_FILL
ws4.cell(row=4, column=2).font = HEADER_FONT
for col in range(1, 3):
    ws4.cell(row=4, column=col).alignment = CENTER
    ws4.cell(row=4, column=col).border = BORDER

for i, (title, body) in enumerate(BMC, 5):
    ws4.cell(row=i, column=1, value=title).font = CAT_FONT
    ws4.cell(row=i, column=1).fill = CAT_FILL
    ws4.cell(row=i, column=1).alignment = Alignment(vertical="top", wrap_text=True)
    ws4.cell(row=i, column=2, value=body).font = NORMAL
    ws4.cell(row=i, column=2).alignment = WRAP
    for col in range(1, 3):
        ws4.cell(row=i, column=col).border = BORDER
    ws4.row_dimensions[i].height = 100

ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 90

# ═══════════════════════════════════════════════════════════
# 저장
# ═══════════════════════════════════════════════════════════
wb.save(OUT_FILE)
print(f"[OK] Generated: {OUT_FILE}")
print(f"   File size: {OUT_FILE.stat().st_size / 1024:.1f} KB")
print(f"   Sheets: 1.ROI-Matrix / 2.100-Ideas / 3.Submission-Form / 4.BMC-Example")
